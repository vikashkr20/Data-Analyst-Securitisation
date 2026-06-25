"""
generate_excel_validations.py
============================
Generates 3 Excel validation workbooks for the Securitisation AI Agent project:
  1. 483553A_ManishKumar_DAX_Dictionary.xlsx
  2. 483553A_ManishKumar_ECL_Validation.xlsx
  3. 483553A_ManishKumar_Waterfall_Validation.xlsx

Usage: python generate_excel_validations.py
"""

import os
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                              numbers)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import SeriesLabel
import warnings
warnings.filterwarnings("ignore")

# ── Paths ─────────────────────────────────────────────────────────────────────
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

LOAN_CSV    = os.path.join(BASE_DIR, "auto_loan_securitisation_data.csv")
DPD_CSV     = os.path.join(BASE_DIR, "dpd_snapshot_history.csv")
DYNAMIC_CSV = os.path.join(BASE_DIR, "dynamic_loss_monthly.csv")
STATIC_CSV  = os.path.join(BASE_DIR, "static_pool_vintage_data.csv")

OUT_DICT      = os.path.join(BASE_DIR, "483553A_ManishKumar_DAX_Dictionary.xlsx")
OUT_ECL       = os.path.join(BASE_DIR, "483553A_ManishKumar_ECL_Validation.xlsx")
OUT_WATERFALL = os.path.join(BASE_DIR, "483553A_ManishKumar_Waterfall_Validation.xlsx")

# ── Style Helpers ──────────────────────────────────────────────────────────────
DARK_BLUE   = "1F3864"
GOLD        = "FFB300"
LIGHT_BLUE  = "D6E4F0"
RED         = "FF4444"
GREEN       = "00C853"
WHITE       = "FFFFFF"
GREY        = "F2F2F2"

def header_style(ws, row=1):
    """Apply dark-blue header style to the given row."""
    fill = PatternFill(fill_type="solid", fgColor=DARK_BLUE)
    font = Font(bold=True, color=WHITE, size=11)
    for cell in ws[row]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def autofit(ws):
    """Auto-fit column widths based on cell content."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 45)

def num_format(ws, col_letter, start_row, end_row, fmt):
    """Apply number format to a column range."""
    for row in range(start_row, end_row + 1):
        ws[f"{col_letter}{row}"].number_format = fmt

def freeze_header(ws):
    ws.freeze_panes = "A2"

def alt_row_fill(ws, data_rows, ncols):
    """Apply alternating row fill for readability."""
    fills = [PatternFill(fill_type="solid", fgColor=WHITE),
             PatternFill(fill_type="solid", fgColor=GREY)]
    for i, row in enumerate(ws.iter_rows(min_row=2, max_row=data_rows + 1,
                                          min_col=1, max_col=ncols)):
        for cell in row:
            cell.fill = fills[i % 2]


# ─────────────────────────────────────────────────────────────────────────────
# WORKBOOK 1: DAX DICTIONARY
# ─────────────────────────────────────────────────────────────────────────────
def build_dax_dictionary():
    print("📖 Building DAX Dictionary workbook...")
    wb = openpyxl.Workbook()

    # ── Sheet 1: Measure Dictionary ──────────────────────────────────────────
    ws = wb.active
    ws.title = "Measure Dictionary"
    ws.row_dimensions[1].height = 30

    headers = ["ID", "Category", "Measure Name", "DAX Formula (Abbreviated)",
               "Business Purpose", "Tables Used", "Data Type", "Format String", "Notes"]
    ws.append(headers)
    header_style(ws)

    measures = [
        # Pool Balance
        (1,"Pool Balance","Total Outstanding Balance","SUM(CurrentBalance) filtered CB>0","Sum of all active outstanding principal","DimLoan/Fact","Currency","₹#,##0.00","Core KPI"),
        (2,"Pool Balance","Original Pool Balance","SUM(OriginalLoanAmount)","Total origination balance","DimLoan","Currency","₹#,##0.00","Denominator for pool factor"),
        (3,"Pool Balance","Active Loan Count","COUNTROWS(FILTER(loans, CB>0))","Number of loans with positive balance","DimLoan","Integer","#,##0","Excludes prepaid/written-off"),
        (4,"Pool Balance","Defaulted Loan Count","CALCULATE(COUNTROWS, IsDefaulted=TRUE)","Count of defaulted loans","DimLoan","Integer","#,##0","IFRS9 Stage 3 indicator"),
        (5,"Pool Balance","Prepaid Loan Count","CALCULATE(COUNTROWS, CB=0, IsDefaulted=FALSE)","Count of fully prepaid loans","DimLoan","Integer","#,##0","Pool attrition analysis"),
        (6,"Pool Balance","Delinquent Balance 30Plus","CALCULATE(SUM(CB), DPD>=30)","Balance of loans 30+ DPD","DimLoan","Currency","₹#,##0.00","Trigger metric numerator"),
        (7,"Pool Balance","Pool Factor","DIVIDE([Outstanding],[Original])","Proportion of original balance remaining","DimLoan","Decimal","0.0000","Pool amortisation tracker"),
        (8,"Pool Balance","Net Loss Total","SUM(NetLoss)","Total net credit loss post-recoveries","DimLoan","Currency","₹#,##0.00","Loss trigger numerator"),
        (9,"Pool Balance","Total EMI Due","SUM(ScheduledPaymentDue)","Total scheduled EMI for billing","DimLoan","Currency","₹#,##0.00","Collection efficiency denominator"),
        (10,"Pool Balance","Current Balance 0DPD","CALCULATE(SUM(CB), DPD=0)","Performing loan balance","DimLoan","Currency","₹#,##0.00","Complement of delinquent balance"),
        # Delinquency
        (11,"Delinquency","Balance_Current","CALCULATE(SUM(CB), Status='Current')","Balance of fully current loans","DimLoan","Currency","₹#,##0.00","Largest bucket expected"),
        (12,"Delinquency","Balance_1to29DPD","CALCULATE(SUM(CB), Status='1-29 DPD')","Early warning delinquency bucket","DimLoan","Currency","₹#,##0.00","SMA-0 equivalent"),
        (13,"Delinquency","Balance_30to59DPD","CALCULATE(SUM(CB), Status='30-59 DPD')","SMA-1 equivalent bucket","DimLoan","Currency","₹#,##0.00","RBI SMA-1 classification"),
        (14,"Delinquency","Balance_60to89DPD","CALCULATE(SUM(CB), Status='60-89 DPD')","SMA-2 equivalent bucket","DimLoan","Currency","₹#,##0.00","RBI SMA-2 classification"),
        (15,"Delinquency","Balance_90to119DPD","CALCULATE(SUM(CB), Status='90-119 DPD')","NPA boundary bucket","DimLoan","Currency","₹#,##0.00","NPA if persists >90 days"),
        (16,"Delinquency","Balance_120Plus","CALCULATE(SUM(CB), Status IN {120+,Default,Repo})","NPA and defaulted balance","DimLoan","Currency","₹#,##0.00","IFRS9 Stage 3 pool"),
        (17,"Delinquency","Count_Current","CALCULATE(COUNTROWS, Status='Current')","Number of fully current loans","DimLoan","Integer","#,##0","Denominator for rates"),
        (18,"Delinquency","Count_30Plus","CALCULATE(COUNTROWS, DPD>=30)","Number of loans 30+ DPD","DimLoan","Integer","#,##0","Delinquency headcount"),
        (19,"Delinquency","Delinquency_30Plus_Rate","DIVIDE([Delinquent Bal 30+],[Outstanding])","30+ DPD balance as % of pool","DimLoan","Percentage","0.00%","Primary trigger metric"),
        (20,"Delinquency","Delinquency_90Plus_Rate","DIVIDE(CALC(SUM,DPD>=90),[Outstanding])","90+ DPD balance as % of pool","DimLoan","Percentage","0.00%","NPA rate equivalent"),
        # Weighted Averages
        (21,"Weighted Avg","WAC","SUMX(CB*Rate)/[Outstanding]","Blended interest rate of pool","DimLoan","Percentage","0.00%","Key pool yield metric"),
        (22,"Weighted Avg","WAM","SUMX(CB*RemainingTerm)/[Outstanding]","Average months to maturity","DimLoan","Decimal","0.00","Pool tenor profile"),
        (23,"Weighted Avg","WALA","SUMX(CB*MonthsOnBook)/[Outstanding]","Average loan seasoning in months","DimLoan","Decimal","0.00","Seasoning indicator"),
        (24,"Weighted Avg","WALT","SUMX(CB*OriginalTerm)/[Outstanding]","Average original loan term","DimLoan","Decimal","0.00","Used with WALA for life"),
        (25,"Weighted Avg","WA_LTV_Origination","SUMX(OLA*LTV_Orig)/[OrigBalance]","WA LTV at loan origination","DimLoan","Percentage","0.00%","Original credit risk"),
        (26,"Weighted Avg","WA_LTV_Current","SUMX(CB*LTV_Current)/[Outstanding]","Current WA loan-to-value ratio","DimLoan","Percentage","0.00%","Collateral coverage"),
        (27,"Weighted Avg","WA_CIBIL_Score_Current","SUMX(CB*CIBIL_Current)/[Outstanding]","WA current CIBIL credit score","DimLoan/DimBorrower","Decimal","0.00","Pool credit quality"),
        (28,"Weighted Avg","WA_DTI_Ratio","SUMX(CB*DTI)/[Outstanding]","WA debt-to-income ratio","DimLoan/DimBorrower","Percentage","0.00%","Borrower affordability"),
        # Time Intelligence
        (29,"Time Intelligence","YTD_NetLoss","CALCULATE(SUM(NetLoss_Month),DATESYTD)","Year-to-date cumulative net loss","FactDynamicLoss/DimDate","Currency","₹#,##0.00","Annual benchmarking"),
        (30,"Time Intelligence","QTD_NetLoss","CALCULATE(SUM(NetLoss_Month),DATESQTD)","Quarter-to-date net loss","FactDynamicLoss/DimDate","Currency","₹#,##0.00","Quarterly reporting"),
        (31,"Time Intelligence","MTD_Collections","CALCULATE(SUM(Collections),DATESMTD)","Month-to-date collections","FactDynamicLoss/DimDate","Currency","₹#,##0.00","In-month tracker"),
        (32,"Time Intelligence","Rolling_3M_DefaultRate","AVERAGEX(DATESINPERIOD(-3M),DefaultRate)","3-month rolling avg default rate","FactDynamicLoss/DimDate","Percentage","0.000%","Smoothed trend"),
        (33,"Time Intelligence","Rolling_12M_AvgDelinquencyRate","AVERAGEX(DATESINPERIOD(-12M),DefRate)","12-month rolling delinquency avg","FactDynamicLoss/DimDate","Percentage","0.000%","Trend deviation baseline"),
        (34,"Time Intelligence","MoM_Balance_Change","[Outstanding]-CALCULATE([Outstanding],PREVIOUSMONTH)","Month-over-month balance change","DimLoan/DimDate","Currency","₹#,##0.00","Amortisation speed"),
        (35,"Time Intelligence","MoM_DefaultRate_Change","[Delinq30+Rate]-CALC([Delinq30+Rate],PREVMONTH)","MoM change in delinquency rate","DimLoan/DimDate","Percentage","0.00%","Early warning delta"),
        (36,"Time Intelligence","YoY_Balance_Change","[Outstanding]-CALC([Outstanding],PARALLELPERIOD-12M)","Year-on-year balance movement","DimLoan/DimDate","Currency","₹#,##0.00","Annual trend"),
        (37,"Time Intelligence","PriorMonth_Balance","CALCULATE([Outstanding],PREVIOUSMONTH)","Prior month outstanding balance","DimLoan/DimDate","Currency","₹#,##0.00","MoM comparison base"),
        (38,"Time Intelligence","StdDev_12M_DelinquencyRate","STDEV.P(MonthlyDefaultRate) last 12M","Standard deviation of default rate","FactDynamicLoss/DimDate","Decimal","0.0000","Volatility measure"),
        # Concentration
        (39,"Concentration","HHI_Geographic_Concentration","SUMX(States,(StateBalance/Total)^2)","Geographic concentration HHI index","DimGeography","Decimal","0.0000","HHI>0.25=high concentration"),
        (40,"Concentration","Top5_State_Balance_Share","SUMX(TOPN(5,States,Balance),Bal)/Total","Top-5 state combined balance share","DimGeography","Percentage","0.00%","Geographic risk"),
        (41,"Concentration","Top10_Borrower_Balance_Share","SUMX(TOPN(10,Borrowers,Bal),Bal)/Total","Top-10 obligor concentration","DimBorrower","Percentage","0.00%","Single-name risk"),
        (42,"Concentration","SingleObligor_MaxExposure","MAXX(SUMMARIZE(Borrowers),BorrowerBal)","Largest single borrower exposure","DimBorrower","Currency","₹#,##0.00","Limit monitoring"),
        (43,"Concentration","VehicleType_Concentration_HHI","SUMX(VehicleTypes,(TypeBal/Total)^2)","Vehicle type concentration HHI","DimVehicle","Decimal","0.0000","Collateral diversity"),
        (44,"Concentration","Servicer_Concentration","MAX(SvcBal)/[Outstanding]","Largest servicer's share of pool","DimLoan","Percentage","0.00%","Operational risk"),
        # Waterfall
        (45,"Waterfall","Waterfall_TotalCollections","SUM(CollectionsTotal)","Total period collections","FactDynamicLoss","Currency","₹#,##0.00","Waterfall input"),
        (46,"Waterfall","Waterfall_Senior_Balance","[Outstanding]*0.75","Senior tranche notional (75% of pool)","DimLoan","Currency","₹#,##0.00","Assumed 75% seniorisation"),
        (47,"Waterfall","Waterfall_Senior_Interest","[SeniorBal]*0.08/12","Monthly senior coupon at 8% p.a.","DimLoan","Currency","₹#,##0.00","Step 1 waterfall"),
        (48,"Waterfall","Waterfall_Senior_Principal","BalanceDrop*0.75","Senior tranche scheduled amortisation","DimLoan/DimDate","Currency","₹#,##0.00","Step 2 waterfall"),
        (49,"Waterfall","Waterfall_Senior_Total","[SeniorInt]+[SeniorPrinc]","Total senior tranche payment","DimLoan/DimDate","Currency","₹#,##0.00","Senior noteholder P&I"),
        (50,"Waterfall","Waterfall_Mezzanine_Balance","[Outstanding]*0.15","Mezzanine tranche notional (15%)","DimLoan","Currency","₹#,##0.00","Assumed 15% mezz"),
        (51,"Waterfall","Waterfall_Mezzanine_Interest","[MezzBal]*0.10/12","Monthly mezzanine coupon at 10% p.a.","DimLoan","Currency","₹#,##0.00","Step 3 waterfall"),
        (52,"Waterfall","Waterfall_Mezzanine_Total","[MezzInt]+[MezzPrinc]","Total mezzanine tranche payment","DimLoan/DimDate","Currency","₹#,##0.00","Mezz noteholder P&I"),
        (53,"Waterfall","Waterfall_ReserveAccount_Contribution","MAX(0,[Outstanding]*0.02*0.05)","Reserve account top-up contribution","DimLoan","Currency","₹#,##0.00","2% reserve target"),
        (54,"Waterfall","Waterfall_Equity_Residual","MAX(0,Collections-Senior-Mezz-Reserve)","Residual equity/subordinate cash flow","FactDynamicLoss","Currency","₹#,##0.00","Equity return"),
        (55,"Waterfall","OC_Ratio_Current","[Outstanding]/(SeniorBal+MezzBal)","Overcollateralisation ratio","DimLoan","Decimal","0.00x","Trigger: breach if <1.05"),
        # Triggers
        (56,"Trigger","Delinquency_Trigger_Proximity","0.08-[Delinq90+Rate]","Bps gap to 8% delinquency trigger","DimLoan","Percentage","0.00%","Positive=safe, Negative=breach"),
        (57,"Trigger","Loss_Trigger_Proximity","0.02-[CumNetLoss/OrigBal]","Gap to 2% cumulative loss trigger","DimLoan","Percentage","0.00%","Positive=safe"),
        (58,"Trigger","OC_Trigger_Breach","IF([OC_Ratio]<1.05,1,0)","OC trigger breach flag","DimLoan","Integer","0;0;[Red]1","1=breach event"),
        (59,"Trigger","Delinquency_Trigger_Breach","IF([Delinq90+Rate]>0.08,1,0)","Delinquency trigger breach flag","DimLoan","Integer","0;0;[Red]1","1=breach event"),
        (60,"Trigger","Loss_Trigger_Breach","IF([CumLossRate]>0.02,1,0)","Loss trigger breach flag","DimLoan","Integer","0;0;[Red]1","1=breach event"),
        (61,"Trigger","ExcessSpread_Current","[WAC]/100-(0.08*0.75+0.10*0.15)-0.01","Net excess spread buffer","DimLoan","Percentage","0.00%","Protection layer"),
        # IFRS9
        (62,"IFRS9 ECL","IFRS9_Stage1_Count","CALCULATE(COUNTROWS, Stage=1)","Count of Stage 1 (performing) loans","DimLoan","Integer","#,##0","12-month ECL bucket"),
        (63,"IFRS9 ECL","IFRS9_Stage2_Count","CALCULATE(COUNTROWS, Stage=2)","Count of Stage 2 (SICR) loans","DimLoan","Integer","#,##0","Lifetime ECL - watch list"),
        (64,"IFRS9 ECL","IFRS9_Stage3_Count","CALCULATE(COUNTROWS, Stage=3)","Count of Stage 3 (impaired) loans","DimLoan","Integer","#,##0","Lifetime ECL - NPA"),
        (65,"IFRS9 ECL","Stage1_ECL_Total","SUMX(Stage1, PD*LGD*EAD)","12-month ECL for Stage 1 loans","DimLoan","Currency","₹#,##0.00","PD=12M point-in-time"),
        (66,"IFRS9 ECL","Stage2_ECL_Total","SUMX(Stage2, PD*1.5*LGD*EAD)","Lifetime ECL for Stage 2 loans","DimLoan","Currency","₹#,##0.00","Lifetime PD uplift 1.5x"),
        (67,"IFRS9 ECL","Stage3_ECL_Total","SUMX(Stage3, PD*2.0*LGD*EAD)","Lifetime ECL for Stage 3 loans","DimLoan","Currency","₹#,##0.00","Full lifetime PD"),
        (68,"IFRS9 ECL","Total_ECL_Provision","Stage1+Stage2+Stage3 ECL","Total IFRS 9 expected credit loss provision","DimLoan","Currency","₹#,##0.00","P&L provision charge"),
        (69,"IFRS9 ECL","ECL_as_Pct_of_Balance","[Total ECL]/[Outstanding]","ECL coverage as % of pool balance","DimLoan","Percentage","0.00%","Provisioning adequacy"),
        (70,"IFRS9 ECL","ECL_Coverage_Ratio","[Total ECL]/[EAD Total]","ECL to EAD coverage ratio","DimLoan","Percentage","0.00%","Regulatory metric"),
        (71,"IFRS9 ECL","PD_Weighted_Average","SUMX(EAD*PD)/[EAD Total]","Portfolio WA probability of default","DimLoan","Percentage","0.000%","Credit quality summary"),
        (72,"IFRS9 ECL","LGD_Weighted_Average","SUMX(EAD*LGD)/[EAD Total]","Portfolio WA loss given default","DimLoan","Percentage","0.000%","Recovery rate complement"),
        (73,"IFRS9 ECL","EAD_Total","SUM(EAD)","Total exposure at default","DimLoan","Currency","₹#,##0.00","ECL denominator"),
        # Stress Testing
        (74,"Stress Testing","Stressed_DefaultRate","[Delinq30+Rate]*DefaultMultiplier","Stress-adjusted default rate","DimLoan","Percentage","0.00%","What-if parameter driven"),
        (75,"Stress Testing","Stressed_LGD","[WA_LGD]+(1-[WA_LGD])*RecoveryHaircut","Stressed loss given default","DimLoan","Percentage","0.000%","Recovery reduction"),
        (76,"Stress Testing","Stressed_ECL_Stage1","SUMX(Stage1, StressedPD*StressedLGD*EAD)","Stressed 12M ECL - Stage 1","DimLoan","Currency","₹#,##0.00","Scenario ECL"),
        (77,"Stress Testing","Stressed_ECL_Stage2","SUMX(Stage2, StressedPD*1.5*StressedLGD*EAD)","Stressed lifetime ECL - Stage 2","DimLoan","Currency","₹#,##0.00","Scenario ECL"),
        (78,"Stress Testing","Stressed_ECL_Stage3","SUMX(Stage3, StressedPD*2.0*StressedLGD*EAD)","Stressed lifetime ECL - Stage 3","DimLoan","Currency","₹#,##0.00","Scenario ECL"),
        (79,"Stress Testing","Stressed_ECL_Total","S1+S2+S3 Stressed ECL","Total stressed ECL provision","DimLoan","Currency","₹#,##0.00","Total scenario impact"),
        (80,"Stress Testing","Stressed_NetLoss","SUM(LossAmount)*(1+(Mult-1)*0.5)*(1+Haircut)","Stressed net loss amount","DimLoan","Currency","₹#,##0.00","P&L stress impact"),
        (81,"Stress Testing","Stressed_OC_Ratio","(Outstanding-StressedLoss)/(Senior+Mezz)","Stressed overcollateralisation ratio","DimLoan","Decimal","0.00x","Trigger proximity in stress"),
        # Vintage
        (82,"Vintage Analysis","CumulativeNetLossRate_ByVintage","AVERAGE(CumulativeNetLossRate)","Avg cumulative net loss rate by vintage","FactStaticPool","Percentage","0.000%","Loss curve Y-axis"),
        (83,"Vintage Analysis","MarginalLossRate_Peak","MAXX(Vintages, MAX(MarginalLossRate))","Peak monthly marginal loss rate","FactStaticPool","Percentage","0.0000%","Loss timing peak"),
        (84,"Vintage Analysis","LossCurve_Projection","CumLossRate+AvgMarginal*(AvgTerm-WALA)","Projected lifetime loss rate","FactStaticPool","Percentage","0.000%","Rating agency submission"),
        (85,"Vintage Analysis","PoolFactor_ByVintage","AVERAGE(PoolFactor)","Pool factor by vintage cohort","FactStaticPool","Decimal","0.0000","Amortisation pace"),
        (86,"Vintage Analysis","WAC_ByVintage","CALCULATE([WAC], USERELATIONSHIP)","WAC at vintage cohort level","DimLoan/DimDate","Percentage","0.00%","Vintage yield profile"),
    ]

    for m in measures:
        ws.append(list(m))

    alt_row_fill(ws, len(measures), len(headers))
    autofit(ws)
    freeze_header(ws)

    # ── Sheet 2: Table Relationships ──────────────────────────────────────────
    ws2 = wb.create_sheet("Table Relationships")
    rel_headers = ["From Table", "From Column", "To Table", "To Column",
                   "Cardinality", "Filter Direction", "Is Active", "Notes"]
    ws2.append(rel_headers)
    header_style(ws2)

    relationships = [
        ("FactMonthlyPerformance","LoanID","DimLoan","LoanID","Many-to-One","Single","Yes","Primary loan dimension link"),
        ("FactMonthlyPerformance","SnapshotDate","DimDate","Date","Many-to-One","Single","Yes","Reporting date dimension"),
        ("FactDynamicLoss","ReportingDate","DimDate","Date","Many-to-One","Single","Yes","Pool-level monthly reporting"),
        ("FactStaticPool","VintageStartDate","DimDate","Date","Many-to-One","Single","Yes","Vintage origination date"),
        ("DimLoan","OriginationDate","DimDate","Date","Many-to-One","Single","No","Role-playing dimension (inactive)"),
        ("DimLoan","MaturityDate","DimDate","Date","Many-to-One","Single","No","Role-playing dimension (inactive)"),
        ("DimLoan","BorrowerID","DimBorrower","BorrowerID","Many-to-One","Single","Yes","Borrower attributes"),
        ("DimLoan","Region","DimGeography","Region","Many-to-One","Single","Yes","Geographic dimension"),
        ("DimLoan","VehicleType","DimVehicle","VehicleType","Many-to-One","Single","Yes","Collateral dimension"),
    ]
    for r in relationships:
        ws2.append(list(r))
    alt_row_fill(ws2, len(relationships), len(rel_headers))
    autofit(ws2)
    freeze_header(ws2)

    # ── Sheet 3: Calculation Groups ───────────────────────────────────────────
    ws3 = wb.create_sheet("Calculation Groups")
    cg_headers = ["Group Name", "Item Name", "Expression", "Ordinal", "Description"]
    ws3.append(cg_headers)
    header_style(ws3)

    calc_groups = [
        ("Time Intelligence","MTD","CALCULATE(SELECTEDMEASURE(), DATESMTD(DimDate[Date]))",1,"Month-to-date accumulation"),
        ("Time Intelligence","QTD","CALCULATE(SELECTEDMEASURE(), DATESQTD(DimDate[Date]))",2,"Quarter-to-date accumulation"),
        ("Time Intelligence","YTD","CALCULATE(SELECTEDMEASURE(), DATESYTD(DimDate[Date]))",3,"Year-to-date accumulation"),
        ("Time Intelligence","Prior Month","CALCULATE(SELECTEDMEASURE(), PREVIOUSMONTH(DimDate[Date]))",4,"Prior month comparison"),
        ("Time Intelligence","Prior Year","CALCULATE(SELECTEDMEASURE(), PARALLELPERIOD(DimDate[Date],-12,MONTH))",5,"Prior year comparison"),
        ("Time Intelligence","MoM Change","SELECTEDMEASURE()-CALCULATE(SELECTEDMEASURE(),PREVIOUSMONTH(DimDate[Date]))",6,"Month-over-month delta"),
        ("Time Intelligence","YoY Change","SELECTEDMEASURE()-CALCULATE(SELECTEDMEASURE(),PARALLELPERIOD(DimDate[Date],-12,MONTH))",7,"Year-over-year delta"),
        ("Stress Scenarios","Base Case","SELECTEDMEASURE()",1,"No stress adjustments applied"),
        ("Stress Scenarios","Moderate Stress","CALCULATE(SELECTEDMEASURE(), 'DefaultRateMultiplier'[Value]=2, 'RecoveryHaircut'[Value]=0.15)",2,"2x default, 15% recovery haircut"),
        ("Stress Scenarios","Severe Stress","CALCULATE(SELECTEDMEASURE(), 'DefaultRateMultiplier'[Value]=3, 'RecoveryHaircut'[Value]=0.30)",3,"3x default, 30% recovery haircut"),
        ("Stress Scenarios","Crisis","CALCULATE(SELECTEDMEASURE(), 'DefaultRateMultiplier'[Value]=5, 'RecoveryHaircut'[Value]=0.50)",4,"5x default, 50% recovery haircut"),
    ]
    for cg in calc_groups:
        ws3.append(list(cg))
    alt_row_fill(ws3, len(calc_groups), len(cg_headers))
    autofit(ws3)
    freeze_header(ws3)

    wb.save(OUT_DICT)
    print(f"  ✅ Saved: {os.path.basename(OUT_DICT)}")


# ─────────────────────────────────────────────────────────────────────────────
# WORKBOOK 2: ECL VALIDATION
# ─────────────────────────────────────────────────────────────────────────────
def build_ecl_validation():
    print("📊 Building ECL Validation workbook...")

    loans = pd.read_csv(LOAN_CSV)
    dpd   = pd.read_csv(DPD_CSV)

    # ── ECL Stage Classification & Calculation ────────────────────────────────
    def classify_stage(row):
        if row['IsDefaulted'] or row['DelinquencyDays'] >= 120:
            return 3
        elif row['DelinquencyDays'] >= 60:
            return 2
        else:
            return 1

    loans['Calc_Stage'] = loans.apply(classify_stage, axis=1)

    def compute_ecl(row):
        pd_est = row['PD_Estimate']
        lgd    = row['LGD_Estimate']
        ead    = row['EAD']
        stage  = row['Calc_Stage']
        if stage == 1:
            return pd_est * lgd * ead
        elif stage == 2:
            return pd_est * 1.5 * lgd * ead
        else:
            return pd_est * 2.0 * lgd * ead

    loans['ECL_Computed'] = loans.apply(compute_ecl, axis=1)
    loans['ECL_Variance'] = loans['ECL_Computed'] - loans['ECL_Provision']
    loans['Variance_Pct'] = (loans['ECL_Variance'] / loans['ECL_Provision'].replace(0, np.nan)).fillna(0)

    wb = openpyxl.Workbook()

    # ── Sheet 1: Loan Level ECL ───────────────────────────────────────────────
    ws = wb.active
    ws.title = "Loan Level ECL"
    ws.row_dimensions[1].height = 30

    ecl_cols = ["LoanID", "DelinquencyStatus", "DelinquencyDays", "CurrentBalance",
                "IFRS9_Stage_Given", "Calc_Stage", "PD_Estimate", "LGD_Estimate",
                "EAD", "ECL_Computed", "ECL_Provision_Given", "Variance_INR", "Variance_Pct"]

    ecl_data = loans[["LoanID","DelinquencyStatus","DelinquencyDays","CurrentBalance",
                       "IFRS9_Stage","Calc_Stage","PD_Estimate","LGD_Estimate",
                       "EAD","ECL_Computed","ECL_Provision","ECL_Variance","Variance_Pct"]].copy()
    ecl_data.columns = ecl_cols

    ws.append(ecl_cols)
    header_style(ws)

    for _, row in ecl_data.iterrows():
        ws.append(list(row))

    # Format numeric columns
    for row in ws.iter_rows(min_row=2, max_row=len(ecl_data)+1):
        row[3].number_format  = '₹#,##0.00'   # CurrentBalance
        row[6].number_format  = '0.000%'        # PD
        row[7].number_format  = '0.000%'        # LGD
        row[8].number_format  = '₹#,##0.00'    # EAD
        row[9].number_format  = '₹#,##0.00'    # ECL Computed
        row[10].number_format = '₹#,##0.00'    # ECL Provision
        row[11].number_format = '₹#,##0.00'    # Variance
        row[12].number_format = '0.00%'          # Variance %

    # Conditional format: Stage 3 rows in light red
    red_fill = PatternFill(fill_type="solid", fgColor="FFCCCC")
    for i, row in enumerate(ws.iter_rows(min_row=2, max_row=len(ecl_data)+1), start=2):
        stage_val = ws.cell(row=i, column=6).value
        if stage_val == 3:
            for cell in ws[i]:
                cell.fill = red_fill
        elif stage_val == 2:
            for cell in ws[i]:
                cell.fill = PatternFill(fill_type="solid", fgColor="FFF3CC")

    autofit(ws)
    freeze_header(ws)

    # ── Sheet 2: Transition Matrix ────────────────────────────────────────────
    ws2 = wb.create_sheet("Transition Matrix")
    ws2.row_dimensions[1].height = 25

    dpd_sorted = dpd.sort_values(['LoanID','SnapshotDate'])
    bucket_order = ['Current','1-29 DPD','30-59 DPD','60-89 DPD','90-119 DPD','120+ DPD','Default']

    transitions = dpd_sorted[dpd_sorted['TransitionType'] != 'Static'][['DPD_Bucket_Prior','DPD_Bucket']].copy()
    transitions = transitions[transitions['DPD_Bucket_Prior'].isin(bucket_order)]
    transitions = transitions[transitions['DPD_Bucket'].isin(bucket_order)]

    matrix_count = pd.crosstab(transitions['DPD_Bucket_Prior'], transitions['DPD_Bucket'])
    # Reindex to ensure all buckets present
    matrix_count = matrix_count.reindex(index=bucket_order, columns=bucket_order, fill_value=0)

    # Write count matrix
    ws2.cell(row=1, column=1, value="TRANSITION MATRIX — COUNT (From \\ To)")
    ws2.cell(row=1, column=1).font = Font(bold=True, size=12)
    ws2.cell(row=2, column=1, value="From \\ To")
    for j, col in enumerate(matrix_count.columns, start=2):
        ws2.cell(row=2, column=j, value=col)
    header_style(ws2, row=2)

    for i, idx in enumerate(matrix_count.index, start=3):
        ws2.cell(row=i, column=1, value=idx)
        ws2.cell(row=i, column=1).font = Font(bold=True)
        for j, col in enumerate(matrix_count.columns, start=2):
            ws2.cell(row=i, column=j, value=int(matrix_count.loc[idx, col]))

    # Write % matrix below
    matrix_pct = matrix_count.div(matrix_count.sum(axis=1).replace(0, np.nan), axis=0).fillna(0)
    start_row = len(matrix_count) + 5

    ws2.cell(row=start_row, column=1, value="TRANSITION MATRIX — % RATE (From \\ To)")
    ws2.cell(row=start_row, column=1).font = Font(bold=True, size=12)
    ws2.cell(row=start_row+1, column=1, value="From \\ To")
    for j, col in enumerate(matrix_pct.columns, start=2):
        ws2.cell(row=start_row+1, column=j, value=col)
    header_style(ws2, row=start_row+1)

    for i, idx in enumerate(matrix_pct.index, start=start_row+2):
        ws2.cell(row=i, column=1, value=idx)
        ws2.cell(row=i, column=1).font = Font(bold=True)
        for j, col in enumerate(matrix_pct.columns, start=2):
            c = ws2.cell(row=i, column=j, value=round(matrix_pct.loc[idx, col], 4))
            c.number_format = '0.00%'
            # Color scale: green for diagonal (cure/same), red for worsening
            val = matrix_pct.loc[idx, col]
            if j - 2 < bucket_order.index(idx) and val > 0.1:
                c.fill = PatternFill(fill_type="solid", fgColor="FFCCCC")
            elif j - 2 == bucket_order.index(idx):
                c.fill = PatternFill(fill_type="solid", fgColor="CCE5FF")

    autofit(ws2)

    # ── Sheet 3: ECL Summary ──────────────────────────────────────────────────
    ws3 = wb.create_sheet("ECL Summary")
    ws3.row_dimensions[1].height = 30

    summary_headers = ["Dimension", "Category", "Loan Count", "Total Balance (₹)",
                       "Total EAD (₹)", "Total ECL (₹)", "ECL Coverage %", "Avg PD %", "Avg LGD %"]
    ws3.append(summary_headers)
    header_style(ws3)

    # By IFRS9 Stage
    for stage in [1, 2, 3]:
        sub = loans[loans['Calc_Stage'] == stage]
        row = [
            "IFRS9 Stage", f"Stage {stage}",
            len(sub),
            sub['CurrentBalance'].sum(),
            sub['EAD'].sum(),
            sub['ECL_Computed'].sum(),
            sub['ECL_Computed'].sum() / sub['EAD'].sum() if sub['EAD'].sum() > 0 else 0,
            sub['PD_Estimate'].mean(),
            sub['LGD_Estimate'].mean()
        ]
        ws3.append(row)

    # Total row
    ws3.append([
        "TOTAL", "All Stages",
        len(loans), loans['CurrentBalance'].sum(), loans['EAD'].sum(),
        loans['ECL_Computed'].sum(),
        loans['ECL_Computed'].sum() / loans['EAD'].sum(),
        loans['PD_Estimate'].mean(), loans['LGD_Estimate'].mean()
    ])

    ws3.append([])

    # By Delinquency Status
    ws3.append(["Dimension","Category","Loan Count","Total Balance (₹)","Total EAD (₹)",
                 "Total ECL (₹)","ECL Coverage %","Avg PD %","Avg LGD %"])
    header_style(ws3, ws3.max_row)

    for status in loans['DelinquencyStatus'].unique():
        sub = loans[loans['DelinquencyStatus'] == status]
        row = [
            "Delinquency Status", status, len(sub),
            sub['CurrentBalance'].sum(), sub['EAD'].sum(), sub['ECL_Computed'].sum(),
            sub['ECL_Computed'].sum() / sub['EAD'].sum() if sub['EAD'].sum() > 0 else 0,
            sub['PD_Estimate'].mean(), sub['LGD_Estimate'].mean()
        ]
        ws3.append(row)

    # Format numeric columns
    for row in ws3.iter_rows(min_row=2, max_row=ws3.max_row):
        for cell in row:
            if isinstance(cell.value, float):
                if cell.column in [4, 5, 6]:
                    cell.number_format = '₹#,##0.00'
                elif cell.column in [7, 8, 9]:
                    cell.number_format = '0.00%'

    autofit(ws3)
    freeze_header(ws3)

    # ── Sheet 4: Stage Migration ───────────────────────────────────────────────
    ws4 = wb.create_sheet("Stage Migration")
    ws4.row_dimensions[1].height = 30

    mig_headers = ["Snapshot Date", "Stage 1 Count", "Stage 2 Count", "Stage 3 Count",
                   "Stage 1 Balance (₹)", "Stage 2 Balance (₹)", "Stage 3 Balance (₹)",
                   "Stage 1 %", "Stage 2 %", "Stage 3 %"]
    ws4.append(mig_headers)
    header_style(ws4)

    def get_stage(dpd_days, is_defaulted=False):
        if is_defaulted or dpd_days >= 120:
            return 3
        elif dpd_days >= 60:
            return 2
        return 1

    for snap_date in sorted(dpd['SnapshotDate'].unique()):
        snap = dpd[dpd['SnapshotDate'] == snap_date].copy()
        snap['Stage'] = snap['DPD_Days'].apply(lambda x: get_stage(x))
        total_bal = snap['CurrentBalance'].sum()
        row = [snap_date]
        for s in [1, 2, 3]:
            sub = snap[snap['Stage'] == s]
            row.append(len(sub))
        for s in [1, 2, 3]:
            sub = snap[snap['Stage'] == s]
            row.append(sub['CurrentBalance'].sum())
        for s in [1, 2, 3]:
            sub = snap[snap['Stage'] == s]
            row.append(sub['CurrentBalance'].sum() / total_bal if total_bal > 0 else 0)
        ws4.append(row)

    for row in ws4.iter_rows(min_row=2, max_row=ws4.max_row):
        for cell in row:
            if cell.column in [5, 6, 7]:
                cell.number_format = '₹#,##0.00'
            elif cell.column in [8, 9, 10]:
                cell.number_format = '0.00%'

    autofit(ws4)
    freeze_header(ws4)

    wb.save(OUT_ECL)
    print(f"  ✅ Saved: {os.path.basename(OUT_ECL)}")


# ─────────────────────────────────────────────────────────────────────────────
# WORKBOOK 3: WATERFALL VALIDATION
# ─────────────────────────────────────────────────────────────────────────────
def build_waterfall_validation():
    print("🏦 Building Waterfall Validation workbook...")

    dyn = pd.read_csv(DYNAMIC_CSV)
    dyn = dyn.sort_values('ReportingDate').reset_index(drop=True)

    # Tranche assumptions
    SENIOR_PCT   = 0.75
    MEZZ_PCT     = 0.15
    EQUITY_PCT   = 0.10
    SENIOR_RATE  = 0.08
    MEZZ_RATE    = 0.10
    RESERVE_PCT  = 0.02
    DELINQ_TRIG  = 0.003   # 0.30%/month ≈ 3.6%/year
    LOSS_TRIG    = 0.002   # 0.20%/month
    OC_TRIG      = 1.05

    dyn['Senior_Balance']       = dyn['BOP_Balance'] * SENIOR_PCT
    dyn['Mezz_Balance']         = dyn['BOP_Balance'] * MEZZ_PCT
    dyn['Equity_Balance']       = dyn['BOP_Balance'] * EQUITY_PCT
    dyn['Senior_Interest']      = dyn['Senior_Balance'] * SENIOR_RATE / 12
    dyn['Mezz_Interest']        = dyn['Mezz_Balance']   * MEZZ_RATE   / 12
    dyn['Senior_Principal']     = (dyn['BOP_Balance'] - dyn['EOP_Balance']) * SENIOR_PCT
    dyn['Mezz_Principal']       = (dyn['BOP_Balance'] - dyn['EOP_Balance']) * MEZZ_PCT
    dyn['Senior_Total']         = dyn['Senior_Interest'] + dyn['Senior_Principal']
    dyn['Mezz_Total']           = dyn['Mezz_Interest']   + dyn['Mezz_Principal']
    dyn['After_Senior_Int']     = dyn['CollectionsTotal'] - dyn['Senior_Interest']
    dyn['After_Senior_Princ']   = dyn['After_Senior_Int'] - dyn['Senior_Principal']
    dyn['After_Mezz_Int']       = dyn['After_Senior_Princ'] - dyn['Mezz_Interest']
    dyn['Reserve_Contribution'] = (dyn['BOP_Balance'] * RESERVE_PCT * 0.05).clip(lower=0)
    dyn['Equity_Residual']      = (dyn['After_Mezz_Int'] - dyn['Reserve_Contribution'] - dyn['Mezz_Principal']).clip(lower=0)
    dyn['OC_Ratio']             = dyn['BOP_Balance'] / (dyn['Senior_Balance'] + dyn['Mezz_Balance'])
    dyn['Excess_Spread_Annual'] = dyn['CollectionsTotal'] / dyn['BOP_Balance'] * 12 - SENIOR_RATE * SENIOR_PCT - MEZZ_RATE * MEZZ_PCT
    dyn['Loss_Trigger_Flag']    = (dyn['MonthlyNetLossRate'] > LOSS_TRIG).astype(int)
    dyn['Delinq_Trigger_Flag']  = (dyn['MonthlyDefaultRate'] > DELINQ_TRIG).astype(int)
    dyn['OC_Trigger_Flag']      = (dyn['OC_Ratio'] < OC_TRIG).astype(int)

    wb = openpyxl.Workbook()

    # ── Sheet 1: Monthly Waterfall ─────────────────────────────────────────────
    ws = wb.active
    ws.title = "Monthly Waterfall"
    ws.row_dimensions[1].height = 30

    wf_cols = [
        "Reporting Date","BOP Balance","EOP Balance","Collections Total",
        "Senior Balance (75%)","Senior Interest (8% pa)","Senior Principal",
        "Mezz Balance (15%)","Mezz Interest (10% pa)","Mezz Principal",
        "After Senior Interest","After Senior Principal","After Mezz Interest",
        "Reserve Contribution (2%)","Equity Residual","OC Ratio",
        "Excess Spread (Annual)","Net Loss This Month","Monthly Default Rate","Monthly Net Loss Rate"
    ]
    ws.append(wf_cols)
    header_style(ws)

    wf_data_cols = [
        "ReportingDate","BOP_Balance","EOP_Balance","CollectionsTotal",
        "Senior_Balance","Senior_Interest","Senior_Principal",
        "Mezz_Balance","Mezz_Interest","Mezz_Principal",
        "After_Senior_Int","After_Senior_Princ","After_Mezz_Int",
        "Reserve_Contribution","Equity_Residual","OC_Ratio",
        "Excess_Spread_Annual","NetLoss_ThisMonth","MonthlyDefaultRate","MonthlyNetLossRate"
    ]

    for _, row in dyn[wf_data_cols].iterrows():
        ws.append(list(row))

    # Number formatting
    currency_cols = [2,3,4,5,6,7,8,9,10,11,12,13,14,15,18]
    pct_cols      = [17,19,20]
    ratio_cols    = [16]

    for row in ws.iter_rows(min_row=2, max_row=len(dyn)+1):
        for cell in row:
            col = cell.column
            if col in currency_cols:
                cell.number_format = '₹#,##0.00'
            elif col in pct_cols:
                cell.number_format = '0.000%'
            elif col in ratio_cols:
                cell.number_format = '0.00"x"'

    autofit(ws)
    freeze_header(ws)

    # ── Sheet 2: Trigger Status ────────────────────────────────────────────────
    ws2 = wb.create_sheet("Trigger Status")
    ws2.row_dimensions[1].height = 30

    trig_headers = [
        "Reporting Date",
        "Monthly Default Rate","Delinq Trigger Threshold","Delinq Trigger Breach?","Delinq Proximity",
        "Monthly Net Loss Rate","Loss Trigger Threshold","Loss Trigger Breach?","Loss Proximity",
        "OC Ratio","OC Trigger Threshold","OC Trigger Breach?","OC Proximity",
        "Breaches This Month"
    ]
    ws2.append(trig_headers)
    header_style(ws2)

    for _, row in dyn.iterrows():
        delinq_breach = "🔴 BREACH" if row['MonthlyDefaultRate'] > DELINQ_TRIG else "🟢 Safe"
        loss_breach   = "🔴 BREACH" if row['MonthlyNetLossRate']  > LOSS_TRIG   else "🟢 Safe"
        oc_breach     = "🔴 BREACH" if row['OC_Ratio'] < OC_TRIG else "🟢 Safe"
        total_breaches = row['Delinq_Trigger_Flag'] + row['Loss_Trigger_Flag'] + row['OC_Trigger_Flag']
        ws2.append([
            row['ReportingDate'],
            row['MonthlyDefaultRate'], DELINQ_TRIG, delinq_breach, DELINQ_TRIG - row['MonthlyDefaultRate'],
            row['MonthlyNetLossRate'],  LOSS_TRIG,   loss_breach,   LOSS_TRIG  - row['MonthlyNetLossRate'],
            row['OC_Ratio'],            OC_TRIG,     oc_breach,     row['OC_Ratio'] - OC_TRIG,
            total_breaches
        ])

    for row in ws2.iter_rows(min_row=2, max_row=len(dyn)+1):
        for cell in row:
            if cell.column in [2, 3, 5, 6, 7, 9]:
                cell.number_format = '0.000%'
            elif cell.column in [10, 11, 12, 13]:
                cell.number_format = '0.00"x"'
            if isinstance(cell.value, str) and "BREACH" in str(cell.value):
                cell.font = Font(bold=True, color=RED)
            elif isinstance(cell.value, str) and "Safe" in str(cell.value):
                cell.font = Font(bold=True, color=GREEN)

    autofit(ws2)
    freeze_header(ws2)

    # ── Sheet 3: Credit Enhancement ───────────────────────────────────────────
    ws3 = wb.create_sheet("Credit Enhancement")
    ws3.row_dimensions[1].height = 30

    ce_headers = [
        "Reporting Date","Senior OC Level (%)","Mezz OC Level (%)","Equity Subordination (%)",
        "Reserve Account (% Bal)","Excess Spread (Annual %)","Total CE Level (%)"
    ]
    ws3.append(ce_headers)
    header_style(ws3)

    for _, row in dyn.iterrows():
        senior_oc  = (row['BOP_Balance'] - row['Senior_Balance']) / row['BOP_Balance']
        mezz_oc    = (row['BOP_Balance'] - row['Senior_Balance'] - row['Mezz_Balance']) / row['BOP_Balance']
        equity_sub = EQUITY_PCT
        reserve    = RESERVE_PCT
        xs         = max(0, row['Excess_Spread_Annual'])
        total_ce   = senior_oc + mezz_oc + equity_sub + reserve
        ws3.append([row['ReportingDate'], senior_oc, mezz_oc, equity_sub, reserve, xs, total_ce])

    for row in ws3.iter_rows(min_row=2, max_row=len(dyn)+1):
        for cell in row:
            if cell.column > 1:
                cell.number_format = '0.00%'

    autofit(ws3)
    freeze_header(ws3)

    # ── Sheet 4: Pool Factor Trend ─────────────────────────────────────────────
    ws4 = wb.create_sheet("Pool Factor Trend")
    ws4.row_dimensions[1].height = 30

    pf_headers = [
        "Reporting Date","BOP Balance","EOP Balance","Pool Factor",
        "Prepayments","Scheduled Amortisation","Net Loss",
        "Monthly Default Rate","CPR Annualised","Excess Spread"
    ]
    ws4.append(pf_headers)
    header_style(ws4)

    orig_balance = dyn['BOP_Balance'].iloc[0]
    for _, row in dyn.iterrows():
        pool_factor = row['EOP_Balance'] / orig_balance
        ws4.append([
            row['ReportingDate'],
            row['BOP_Balance'], row['EOP_Balance'], pool_factor,
            row['Prepayments_ThisMonth'], row['ScheduledAmort'],
            row['NetLoss_ThisMonth'], row['MonthlyDefaultRate'],
            row['CPR_Annualised'], row['ExcessSpread_Monthly']
        ])

    for row in ws4.iter_rows(min_row=2, max_row=len(dyn)+1):
        for cell in row:
            if cell.column in [2, 3, 5, 6, 7, 10]:
                cell.number_format = '₹#,##0.00'
            elif cell.column == 4:
                cell.number_format = '0.0000'
            elif cell.column in [8, 9]:
                cell.number_format = '0.000%'

    autofit(ws4)
    freeze_header(ws4)

    wb.save(OUT_WATERFALL)
    print(f"  ✅ Saved: {os.path.basename(OUT_WATERFALL)}")


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    print("=" * 60)
    print("  Securitisation AI Agent — Excel Validation Generator")
    print("  Project Code: 483553A | Candidate: Manish Kumar")
    print("=" * 60)
    print()

    build_dax_dictionary()
    build_ecl_validation()
    build_waterfall_validation()

    print()
    print("=" * 60)
    print("✅ All 3 Excel workbooks generated successfully!")
    print(f"   Location: {BASE_DIR}")
    print("=" * 60)
